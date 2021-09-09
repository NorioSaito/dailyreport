from datetime import datetime as dt
import json
import openpyxl
from logging import getLogger, config

from Common import util, mail, constants

with open('log_config.json', encoding='utf-8') as f:
        log_config = json.load(f)

config.dictConfig(log_config)
logger = getLogger(__name__)
logger.info('message')

def report(values):
    # Excel に勤務情報書き込み
    result = _write_excel(values)
    if result:
        print(result, "Excel Error")
        return result
    
    # 日報送信
    print('日報送信')
    result = mail.send_mail(values)
    if result:
        print("Mail Error")
        return result

    return result
    
def _write_excel(values):
    settings = util.get_excel_settings()
    file_path = settings['file_path']
    error = ''

    try:
        # Excel 読み込み
        wb = openpyxl.load_workbook(file_path)
        ws = wb['勤務表']

        # 勤務日から Excel 上の行を算出
        work_day = values['work_day']
        work_date = int(work_day.split('/')[-1])
        row = str(work_date + 12)

        # 作業場所を記録
        cell = 'F' + row
        ws[cell].value = '自宅'

        # 開始時刻を記録
        work_start_time = values['work_start_time']
        hh = work_start_time.split(':')[0]
        mm = work_start_time.split(':')[-1]
        # 時刻を記録
        cell = 'I' + row
        ws[cell].value = hh
        # 分を記録
        cell = 'J' + row
        ws[cell].value = mm

        # 終了時刻を記録
        work_start_time = values['work_end_time']
        hh = work_start_time.split(':')[0]
        mm = work_start_time.split(':')[-1]
        # 時刻を記録
        cell = 'K' + row
        ws[cell].value = hh
        # 分を記録
        cell = 'L' + row
        ws[cell].value = mm

        # 休憩時間を記録
        work_start_time = values['breakout_time']
        hh = work_start_time.split(':')[0]
        mm = work_start_time.split(':')[-1]
        # 時刻を記録
        cell = 'M' + row
        ws[cell].value = hh
        # 分を記録
        cell = 'N' + row
        ws[cell].value = mm

        wb.save(file_path)
    except Exception as e:
        logger.error(e)
        error = constants.EXCEL_ERROR
    
    return error
